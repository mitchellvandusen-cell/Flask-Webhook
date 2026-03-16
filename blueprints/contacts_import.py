# blueprints/contacts_import.py — Contact import from CSV/Excel/TXT files
#
# Upload → parse → column mapping → GHL contact creation via contacts.write scope
# Background import via RQ job for large files (up to 10k contacts)

import os
import io
import csv
import json
import re
import logging
import uuid
from datetime import datetime, timezone

from flask import Blueprint, request, jsonify
from flask_login import login_required, current_user

from db import get_db_connection, return_db_connection

logger = logging.getLogger(__name__)

contacts_import_bp = Blueprint('contacts_import', __name__)

# Max file size: 10MB
MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_ROWS = 10000
PREVIEW_ROWS = 10

# GHL contact fields available for mapping
GHL_FIELDS = [
    {"key": "firstName", "label": "First Name"},
    {"key": "lastName", "label": "Last Name"},
    {"key": "phone", "label": "Phone"},
    {"key": "email", "label": "Email"},
    {"key": "address1", "label": "Address"},
    {"key": "city", "label": "City"},
    {"key": "state", "label": "State"},
    {"key": "postalCode", "label": "Zip Code"},
    {"key": "country", "label": "Country"},
    {"key": "companyName", "label": "Company"},
    {"key": "website", "label": "Website"},
    {"key": "source", "label": "Lead Source"},
    {"key": "dateOfBirth", "label": "Date of Birth"},
    {"key": "tags", "label": "Tags"},
    {"key": "notes", "label": "Notes"},
]

# Auto-mapping: common CSV column names → GHL field keys
_AUTO_MAP = {
    "first name": "firstName", "first_name": "firstName", "firstname": "firstName",
    "fname": "firstName", "first": "firstName",
    "last name": "lastName", "last_name": "lastName", "lastname": "lastName",
    "lname": "lastName", "last": "lastName", "surname": "lastName",
    "phone": "phone", "phone number": "phone", "phone_number": "phone",
    "mobile": "phone", "cell": "phone", "cell phone": "phone", "telephone": "phone",
    "email": "email", "email address": "email", "email_address": "email",
    "e-mail": "email",
    "address": "address1", "street": "address1", "address1": "address1",
    "street address": "address1", "address line 1": "address1",
    "city": "city", "town": "city",
    "state": "state", "province": "state", "region": "state",
    "zip": "postalCode", "zip code": "postalCode", "zipcode": "postalCode",
    "postal code": "postalCode", "postal_code": "postalCode", "postcode": "postalCode",
    "country": "country",
    "company": "companyName", "company name": "companyName", "company_name": "companyName",
    "organization": "companyName", "business": "companyName",
    "website": "website", "url": "website", "web": "website",
    "source": "source", "lead source": "source", "lead_source": "source",
    "dob": "dateOfBirth", "date of birth": "dateOfBirth", "birthday": "dateOfBirth",
    "birth date": "dateOfBirth", "date_of_birth": "dateOfBirth",
    "tags": "tags", "tag": "tags", "labels": "tags",
    "notes": "notes", "note": "notes", "comments": "notes", "comment": "notes",
}


def _normalize_phone(raw):
    """Strip formatting from phone number, return digits-only with +1 for US."""
    if not raw:
        return None
    digits = re.sub(r'[^\d]', '', str(raw))
    if not digits:
        return None
    # Remove leading 1 for 11-digit US numbers
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f"+1{digits}"
    # International numbers — return with + prefix
    if len(digits) > 10:
        return f"+{digits}"
    return None  # Too short to be valid


def _parse_csv_data(file_content, filename):
    """Parse CSV or TXT file content. Returns (headers, rows, error)."""
    try:
        text = file_content.decode('utf-8-sig')  # Handle BOM
    except UnicodeDecodeError:
        try:
            text = file_content.decode('latin-1')
        except UnicodeDecodeError:
            return None, None, "Could not decode file. Please use UTF-8 encoding."

    # Detect delimiter
    try:
        sample = text[:8192]
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t|;')
    except csv.Error:
        dialect = csv.excel  # Default to comma

    reader = csv.reader(io.StringIO(text), dialect)
    rows = []
    headers = None
    for i, row in enumerate(reader):
        if i == 0:
            headers = [h.strip() for h in row]
            continue
        if i > MAX_ROWS:
            break
        rows.append(row)

    if not headers:
        return None, None, "File appears to be empty or has no header row."

    return headers, rows, None


def _parse_excel_data(file_content):
    """Parse Excel (.xlsx) file content. Returns (headers, rows, error)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else f"Column {j+1}" for j, c in enumerate(row)]
                continue
            if i > MAX_ROWS:
                break
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        if not headers:
            return None, None, "Excel file appears to be empty."
        return headers, rows, None
    except Exception as e:
        return None, None, f"Failed to parse Excel file: {str(e)}"


def _auto_map_columns(headers):
    """Auto-map detected column names to GHL fields."""
    mapping = {}
    for h in headers:
        normalized = h.lower().strip()
        if normalized in _AUTO_MAP:
            mapping[h] = _AUTO_MAP[normalized]
    return mapping


@contacts_import_bp.route('/api/contacts/upload', methods=['POST'])
@login_required
def upload_file():
    """Upload a file for contact import. Parses headers and returns preview."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    # Check file size
    f.seek(0, 2)
    size = f.tell()
    f.seek(0)
    if size > MAX_FILE_SIZE:
        return jsonify({"error": f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB."}), 400

    filename = f.filename.lower()
    file_content = f.read()

    # Parse based on extension
    if filename.endswith(('.xlsx', '.xls')):
        headers, rows, error = _parse_excel_data(file_content)
    elif filename.endswith(('.csv', '.txt', '.tsv')):
        headers, rows, error = _parse_csv_data(file_content, filename)
    else:
        return jsonify({"error": "Unsupported file type. Use CSV, Excel (.xlsx), or TXT."}), 400

    if error:
        return jsonify({"error": error}), 400

    if not rows:
        return jsonify({"error": "File has headers but no data rows."}), 400

    # Auto-map columns
    auto_mapping = _auto_map_columns(headers)

    # Preview data (first N rows)
    preview = []
    for row in rows[:PREVIEW_ROWS]:
        row_dict = {}
        for j, h in enumerate(headers):
            row_dict[h] = row[j] if j < len(row) else ""
        preview.append(row_dict)

    # Store all rows as JSON for the background job
    all_rows_data = []
    for row in rows:
        row_dict = {}
        for j, h in enumerate(headers):
            row_dict[h] = row[j] if j < len(row) else ""
        all_rows_data.append(row_dict)

    location_id = current_user.location_id
    import_id = str(uuid.uuid4())

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO contact_imports (id, location_id, filename, status, total_rows, preview_data, file_data, created_by)
            VALUES (%s, %s, %s, 'mapping', %s, %s, %s, %s)
        """, (import_id, location_id, f.filename, len(rows),
              json.dumps(preview), json.dumps(all_rows_data), current_user.email))
        conn.commit()
    finally:
        return_db_connection(conn)

    return jsonify({
        "import_id": import_id,
        "filename": f.filename,
        "total_rows": len(rows),
        "headers": headers,
        "auto_mapping": auto_mapping,
        "preview": preview,
        "ghl_fields": GHL_FIELDS,
    })


@contacts_import_bp.route('/api/contacts/import/<import_id>/start', methods=['POST'])
@login_required
def start_import(import_id):
    """Start the import job with column mapping and options."""
    data = request.get_json(silent=True) or {}
    column_mapping = data.get('column_mapping', {})
    duplicate_strategy = data.get('duplicate_strategy', 'skip')
    apply_tags = data.get('apply_tags', [])

    if not column_mapping:
        return jsonify({"error": "No column mapping provided"}), 400

    # Validate at least phone or email is mapped
    mapped_fields = set(column_mapping.values())
    if 'phone' not in mapped_fields and 'email' not in mapped_fields:
        return jsonify({"error": "You must map at least Phone or Email to import contacts."}), 400

    location_id = current_user.location_id

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE contact_imports
            SET column_mapping = %s, duplicate_strategy = %s, apply_tags = %s, status = 'queued'
            WHERE id = %s AND location_id = %s
        """, (json.dumps(column_mapping), duplicate_strategy, json.dumps(apply_tags),
              import_id, location_id))
        conn.commit()
    finally:
        return_db_connection(conn)

    # Queue the background job
    try:
        from extensions import ensure_redis, q_website
        if ensure_redis() and q_website:
            from tasks import import_contacts_task
            q_website.enqueue(
                import_contacts_task,
                import_id,
                job_id=f"import-{import_id[:8]}",
                job_timeout=1800,  # 30 min max
            )
            logger.info(f"Queued contact import job: {import_id}")
        else:
            logger.error("Redis not available for import job")
            return jsonify({"error": "Background job system unavailable. Please try again."}), 503
    except Exception as e:
        logger.error(f"Failed to queue import job: {e}")
        return jsonify({"error": "Failed to start import. Please try again."}), 500

    return jsonify({"ok": True, "import_id": import_id, "status": "queued"})


@contacts_import_bp.route('/api/contacts/import/<import_id>/status', methods=['GET'])
@login_required
def import_status(import_id):
    """Poll import progress."""
    location_id = current_user.location_id
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT status, total_rows, imported, updated, skipped, failed, error_log, completed_at
            FROM contact_imports
            WHERE id = %s AND location_id = %s
        """, (import_id, location_id))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "Import not found"}), 404
        return jsonify({
            "status": row[0],
            "total_rows": row[1],
            "imported": row[2],
            "updated": row[3],
            "skipped": row[4],
            "failed": row[5],
            "error_count": len(row[6]) if row[6] else 0,
            "completed_at": row[7].isoformat() if row[7] else None,
        })
    finally:
        return_db_connection(conn)


@contacts_import_bp.route('/api/contacts/imports', methods=['GET'])
@login_required
def list_imports():
    """List recent imports for the current location."""
    location_id = current_user.location_id
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, filename, status, total_rows, imported, updated, skipped, failed, created_at, completed_at
            FROM contact_imports
            WHERE location_id = %s
            ORDER BY created_at DESC
            LIMIT 20
        """, (location_id,))
        imports = []
        for row in cur.fetchall():
            imports.append({
                "id": row[0], "filename": row[1], "status": row[2],
                "total_rows": row[3], "imported": row[4], "updated": row[5],
                "skipped": row[6], "failed": row[7],
                "created_at": row[8].isoformat() if row[8] else None,
                "completed_at": row[9].isoformat() if row[9] else None,
            })
        return jsonify({"imports": imports})
    finally:
        return_db_connection(conn)


@contacts_import_bp.route('/api/contacts/import/<import_id>/errors', methods=['GET'])
@login_required
def import_errors(import_id):
    """Get detailed error log for an import."""
    location_id = current_user.location_id
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT error_log FROM contact_imports
            WHERE id = %s AND location_id = %s
        """, (import_id, location_id))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "Import not found"}), 404
        return jsonify({"errors": row[0] or []})
    finally:
        return_db_connection(conn)
